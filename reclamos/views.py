from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q
from cuentas.models import Perfil, AsignacionVendedor
from cuentas.views import role_required
from .models import (
    Reclamo, Evidencia, Respuesta, 
    CategoriaProblema, TipoProblema, MotivoRechazo,
    AsignacionArea, HistorialEstado
)
from .forms import ReclamoForm, RespuestaForm
import json
from .notificaciones import (
    notificar_reclamo_creado,
    notificar_reclamo_aprobado,
    notificar_reclamo_rechazado,
    notificar_nueva_respuesta,
    notificar_reclamo_resuelto,
)


# =============================================
# DASHBOARD DEL CLIENTE
# =============================================
@role_required('cliente_comprador')
def dashboard_cliente(request):
    perfil = request.user.perfil
    reclamos = Reclamo.objects.filter(cliente=perfil).order_by('-created_at')

    pendientes = reclamos.filter(estado__in=['pendiente_validacion', 'en_validacion', 'aprobado_pendiente']).count()
    en_proceso = reclamos.filter(estado__in=['en_proceso', 'en_investigacion', 'resolucion_propuesta']).count()
    finalizados = reclamos.filter(estado__in=['resuelto', 'cerrado', 'rechazado']).count()

    return render(request, 'reclamos/dashboard_cliente.html', {
        'perfil': perfil,
        'reclamos': reclamos,
        'pendientes': pendientes,
        'en_proceso': en_proceso,
        'finalizados': finalizados,
    })


# =============================================
# CREAR RECLAMO
# =============================================
@role_required('cliente_comprador')
def crear_reclamo(request):
    perfil = request.user.perfil

    if request.method == 'POST':
        form = ReclamoForm(request.POST, request.FILES)
        if form.is_valid():
            reclamo = form.save(commit=False)
            reclamo.cliente = perfil
            reclamo.save()

            archivos = form.cleaned_data.get('archivos')
            if archivos:
                for archivo in archivos:
                    Evidencia.objects.create(reclamo=reclamo, archivo=archivo)
            notificar_reclamo_creado(reclamo)  # ← AGREGAR
            messages.success(request, 'reclamo_creado')
            
            return redirect('detalle_reclamo', pk=reclamo.pk)
    else:
        form = ReclamoForm()

    return render(request, 'reclamos/crear_reclamo.html', {
        'form': form,
        'perfil': perfil,
    })

@role_required('cliente_comprador', 'empleado')
def detalle_reclamo(request, pk):
    perfil = request.user.perfil
    reclamo = get_object_or_404(Reclamo, pk=pk)

    if perfil.role == 'cliente_comprador' and reclamo.cliente != perfil:
        messages.error(request, 'sin_permiso')
        return redirect('dashboard_cliente')

    if perfil.role == 'empleado' and perfil.departamento not in ['calidad', 'gerencia']:
        if not AsignacionArea.objects.filter(reclamo=reclamo, responsable=perfil).exists():
            messages.error(request, 'sin_permiso')
            return redirect('panel_empleado')

    # ---> ACCIÓN: CLIENTE CONFIRMA CIERRE <---
    if request.method == 'POST' and 'confirmar_cierre' in request.POST:
        if perfil.role == 'cliente_comprador' and reclamo.estado == 'resuelto':
            reclamo.estado = 'cerrado'
            reclamo.save(update_fields=['estado', 'updated_at'])
            HistorialEstado.objects.create(
                reclamo=reclamo, estado_anterior='resuelto',
                estado_nuevo='cerrado', usuario=perfil
            )
            messages.success(request, 'reclamo_cerrado')
            return redirect('detalle_reclamo', pk=reclamo.pk)

    chat_externo = reclamo.respuestas.filter(es_interno=False).select_related('autor__user')
    chat_interno = None
    
    if perfil.role == 'empleado' and perfil.departamento in ['calidad', 'poscosecha', 'produccion', 'exportaciones']:
        chat_interno = reclamo.respuestas.filter(es_interno=True).select_related('autor__user')

    evidencias = reclamo.evidencias.all()
    form = None

    if perfil.departamento == 'calidad':
        
        if request.method == 'POST' and 'aprobar_reclamo' in request.POST:
            estado_anterior = reclamo.estado
            reclamo.estado = 'en_proceso'
            reclamo.inspector_calidad = perfil
            reclamo.save(update_fields=['estado', 'inspector_calidad', 'updated_at'])
            
            HistorialEstado.objects.create(
                reclamo=reclamo, estado_anterior=estado_anterior, 
                estado_nuevo='en_proceso', usuario=perfil
            )

            empleado_elegido_id = request.POST.get('empleado_asignado')
            from cuentas.models import Perfil as PerfilCuenta
            
            try:
                empleado_elegido = PerfilCuenta.objects.get(id=empleado_elegido_id)
                AsignacionArea.objects.create(
                    reclamo=reclamo,
                    responsable=empleado_elegido,
                    departamento=empleado_elegido.get_departamento_display()
                )
                notificar_reclamo_aprobado(reclamo, empleado_elegido)
                messages.success(request, 'reclamo_asignado')
            except PerfilCuenta.DoesNotExist:
                messages.error(request, 'error_asignacion')

            return redirect('detalle_reclamo', pk=reclamo.pk)

        if request.method == 'POST' and 'rechazar_reclamo' in request.POST:
            estado_anterior = reclamo.estado
            motivo_id = request.POST.get('motivo_rechazo')
            obs = request.POST.get('observaciones_calidad', '')
            
            reclamo.estado = 'rechazado'
            reclamo.inspector_calidad = perfil
            reclamo.observaciones_internas_calidad = obs
            if motivo_id:
                reclamo.motivo_rechazo_id = motivo_id
            reclamo.save()
            
            HistorialEstado.objects.create(
                reclamo=reclamo, estado_anterior=estado_anterior, 
                estado_nuevo='rechazado', usuario=perfil, comentarios=obs
            )
            notificar_reclamo_rechazado(reclamo)
            messages.error(request, 'reclamo_rechazado')
            return redirect('detalle_reclamo', pk=reclamo.pk)

        # ---> ACCIÓN: MARCAR COMO RESUELTO <---
        if request.method == 'POST' and 'marcar_resuelto' in request.POST:
            estado_anterior = reclamo.estado
            reclamo.estado = 'resuelto'
            reclamo.save(update_fields=['estado', 'updated_at'])
            HistorialEstado.objects.create(
                reclamo=reclamo, estado_anterior=estado_anterior,
                estado_nuevo='resuelto', usuario=perfil
            )
            notificar_reclamo_resuelto(reclamo)
            messages.success(request, 'reclamo_resuelto')
            return redirect('detalle_reclamo', pk=reclamo.pk)

    elif perfil.departamento in ['poscosecha', 'produccion', 'exportaciones', 'ventas']:

        if request.method == 'POST' and 'asignar_empleado' in request.POST:
            from cuentas.models import Perfil as PerfilCuenta
            empleado_id = request.POST.get('empleado_asignado')
            try:
                empleado = PerfilCuenta.objects.get(id=empleado_id)
                AsignacionArea.objects.create(
                    reclamo=reclamo,
                    responsable=empleado,
                    departamento=empleado.get_departamento_display()
                )
                messages.success(request, 'empleado_asignado')
            except PerfilCuenta.DoesNotExist:
                messages.error(request, 'error_asignacion')
            return redirect('detalle_reclamo', pk=reclamo.pk)

        elif request.method == 'POST':
            form = RespuestaForm(request.POST)
            if form.is_valid():
                respuesta = form.save(commit=False)
                respuesta.reclamo = reclamo
                respuesta.autor = perfil
                respuesta.save()

                if reclamo.estado == 'en_proceso':
                    reclamo.estado = 'resolucion_propuesta'
                    reclamo.save(update_fields=['estado'])
                    
                    HistorialEstado.objects.create(
                        reclamo=reclamo, estado_anterior='en_proceso', 
                        estado_nuevo='resolucion_propuesta', usuario=perfil
                    )

                messages.success(request, 'respuesta_enviada')
                notificar_nueva_respuesta(reclamo, respuesta)
                return redirect('detalle_reclamo', pk=reclamo.pk)
        else:
            form = RespuestaForm()

    motivos_rechazo = MotivoRechazo.objects.filter(activo=True) if perfil.departamento == 'calidad' else None

    empleados_disponibles = None
    if perfil.departamento == 'calidad' and reclamo.estado == 'pendiente_validacion':
        from cuentas.models import Perfil as PerfilCuenta
        empleados_disponibles = PerfilCuenta.objects.filter(
            departamento__in=['poscosecha', 'produccion', 'exportaciones'],
            role='empleado',
            user__is_active=True
        )

    elif perfil.departamento == 'poscosecha' and reclamo.estado in ['en_proceso', 'en_investigacion']:
        from cuentas.models import Perfil as PerfilCuenta
        empleados_disponibles = PerfilCuenta.objects.filter(
            departamento='poscosecha',
            role='empleado',
            user__is_active=True
        )

    return render(request, 'reclamos/detalle_reclamo.html', {
        'reclamo': reclamo,
        'chat_externo': chat_externo,
        'chat_interno': chat_interno,
        'evidencias': evidencias,
        'form': form,
        'perfil': perfil,
        'motivos_rechazo': motivos_rechazo,
        'empleados_disponibles': empleados_disponibles,
    })
# =============================================
# PANEL DE CONTROL DE CALIDAD
# =============================================
@role_required('empleado')
def panel_calidad(request):
    perfil = request.user.perfil
    if perfil.departamento != 'calidad':
        messages.error(request, 'sin_permiso')
        return redirect('login')

    # Reclamos pendientes de validación
    reclamos_pendientes = Reclamo.objects.filter(
        estado__in=['pendiente_validacion', 'en_validacion']
    ).select_related('cliente__user', 'tipo_problema__categoria').order_by('-created_at')

    # Reclamos listos para marcar como resueltos
    reclamos_por_resolver = Reclamo.objects.filter(
        estado='resolucion_propuesta'
    ).select_related('cliente__user', 'tipo_problema__categoria').order_by('-created_at')

    return render(request, 'reclamos/panel_calidad.html', {
        'perfil': perfil,
        'reclamos': reclamos_pendientes,
        'en_calidad_count': reclamos_pendientes.count(),
        'reclamos_por_resolver': reclamos_por_resolver,
        'por_resolver_count': reclamos_por_resolver.count(),
    })


# =============================================
# PANEL GENÉRICO PARA EMPLEADOS
# =============================================
@role_required('empleado')
def panel_empleado(request):
    perfil = request.user.perfil
    
    asignaciones = AsignacionArea.objects.filter(
        responsable=perfil
    ).select_related('reclamo__cliente__user').order_by('-reclamo__created_at')

    pendientes = asignaciones.filter(reclamo__estado='en_proceso').count()

    return render(request, 'reclamos/panel_empleado.html', {
        'perfil': perfil,
        'asignaciones': asignaciones,
        'pendientes': pendientes,
    })


# =============================================
# PANEL POSCOSECHA
# =============================================
@role_required('empleado')
def panel_poscosecha(request):
    perfil = request.user.perfil
    if perfil.departamento != 'poscosecha':
        messages.error(request, 'sin_permiso')
        return redirect('login')

    asignaciones = AsignacionArea.objects.filter(
        responsable=perfil
    ).select_related('reclamo__cliente__user').order_by('-reclamo__created_at')

    pendientes = asignaciones.filter(reclamo__estado='en_proceso').count()

    return render(request, 'reclamos/panel_poscosecha.html', {
        'perfil': perfil,
        'asignaciones': asignaciones,
        'pendientes': pendientes,
    })


# =============================================
# PANEL PRODUCCIÓN
# =============================================
@role_required('empleado')
def panel_produccion(request):
    perfil = request.user.perfil
    if perfil.departamento != 'produccion':
        messages.error(request, 'sin_permiso')
        return redirect('login')

    asignaciones = AsignacionArea.objects.filter(
        responsable=perfil
    ).select_related('reclamo__cliente__user').order_by('-reclamo__created_at')

    pendientes = asignaciones.filter(reclamo__estado='en_proceso').count()

    return render(request, 'reclamos/panel_produccion.html', {
        'perfil': perfil,
        'asignaciones': asignaciones,
        'pendientes': pendientes,
    })


# =============================================
# PANEL EXPORTACIONES
# =============================================
@role_required('empleado')
def panel_exportaciones(request):
    perfil = request.user.perfil
    if perfil.departamento != 'exportaciones':
        messages.error(request, 'sin_permiso')
        return redirect('login')

    asignaciones = AsignacionArea.objects.filter(
        responsable=perfil
    ).select_related('reclamo__cliente__user').order_by('-reclamo__created_at')

    pendientes = asignaciones.filter(reclamo__estado='en_proceso').count()

    return render(request, 'reclamos/panel_exportaciones.html', {
        'perfil': perfil,
        'asignaciones': asignaciones,
        'pendientes': pendientes,
    })


# =============================================
# PANEL GERENCIA
# =============================================
@login_required
def panel_gerencia(request):
    if not request.user.is_superuser:
        return redirect('login')

    reclamos = Reclamo.objects.select_related(
        'cliente__user', 'inspector_calidad__user', 'tipo_problema'
    ).prefetch_related('areas_asignadas__responsable__user').order_by('-created_at')

    total_reclamos = reclamos.count()
    pendientes_calidad = reclamos.filter(estado='pendiente_validacion').count()
    resueltos = reclamos.filter(estado__in=['resuelto', 'cerrado']).count()

    estados_data = Reclamo.objects.values('estado').annotate(total=Count('id'))
    estados_labels = [item['estado'].replace('_', ' ').title() for item in estados_data]
    estados_counts = [item['total'] for item in estados_data]

    areas_data = AsignacionArea.objects.values('departamento').annotate(total=Count('reclamo', distinct=True))
    areas_labels = [item['departamento'] for item in areas_data]
    areas_counts = [item['total'] for item in areas_data]

    return render(request, 'reclamos/panel_gerencia.html', {
        'reclamos': reclamos,
        'total_reclamos': total_reclamos,
        'pendientes_calidad': pendientes_calidad,
        'resueltos': resueltos,
        'estados_labels': json.dumps(estados_labels),
        'estados_counts': json.dumps(estados_counts),
        'areas_labels': json.dumps(areas_labels),
        'areas_counts': json.dumps(areas_counts),
    })

from django.db.models import Avg, F, ExpressionWrapper, DurationField
@login_required
def panel_gerencia(request):
    if not request.user.is_superuser:
        return redirect('login')

    from django.utils import timezone
    from datetime import datetime, timedelta

    # ==========================================
    # FILTROS DE FECHA
    # ==========================================
    periodo = request.GET.get('periodo', 'todo')
    fecha_inicio_str = request.GET.get('fecha_inicio', '')
    fecha_fin_str = request.GET.get('fecha_fin', '')

    hoy = timezone.now()
    fecha_inicio = None
    fecha_fin = None

    if periodo == 'hoy':
        fecha_inicio = hoy.replace(hour=0, minute=0, second=0)
        fecha_fin = hoy
    elif periodo == 'semana':
        fecha_inicio = hoy - timedelta(days=7)
        fecha_fin = hoy
    elif periodo == 'mes':
        fecha_inicio = hoy - timedelta(days=30)
        fecha_fin = hoy
    elif periodo == 'trimestre':
        fecha_inicio = hoy - timedelta(days=90)
        fecha_fin = hoy
    elif periodo == 'anio':
        fecha_inicio = hoy - timedelta(days=365)
        fecha_fin = hoy
    elif periodo == 'personalizado':
        try:
            if fecha_inicio_str:
                fecha_inicio = timezone.make_aware(datetime.strptime(fecha_inicio_str, '%Y-%m-%d'))
            if fecha_fin_str:
                fecha_fin = timezone.make_aware(datetime.strptime(fecha_fin_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59))
        except ValueError:
            fecha_inicio = None
            fecha_fin = None

    # Base queryset
    reclamos = Reclamo.objects.select_related(
        'cliente__user', 'inspector_calidad__user', 'tipo_problema'
    ).prefetch_related('areas_asignadas__responsable__user').order_by('-created_at')

    # Aplicar filtro de fecha
    if fecha_inicio:
        reclamos = reclamos.filter(created_at__gte=fecha_inicio)
    if fecha_fin:
        reclamos = reclamos.filter(created_at__lte=fecha_fin)

    # ==========================================
    # KPIs
    # ==========================================
    total_reclamos = reclamos.count()
    pendientes_calidad = reclamos.filter(estado='pendiente_validacion').count()
    resueltos = reclamos.filter(estado__in=['resuelto', 'cerrado']).count()
    rechazados = reclamos.filter(estado='rechazado').count()

    # Tiempo promedio de resolución en días
    reclamos_resueltos = reclamos.filter(estado__in=['resuelto', 'cerrado'])
    tiempo_promedio_dias = None
    if reclamos_resueltos.exists():
        total_dias = 0
        count = 0
        for r in reclamos_resueltos:
            if r.updated_at and r.created_at:
                diferencia = r.updated_at - r.created_at
                total_dias += diferencia.total_seconds() / 86400
                count += 1
        if count > 0:
            tiempo_promedio_dias = round(total_dias / count, 1)

    # Tasa de aprobación
    validados = reclamos.filter(estado__in=['en_proceso', 'resolucion_propuesta', 'resuelto', 'cerrado', 'rechazado']).count()
    tasa_aprobacion = None
    if validados > 0:
        aprobados = validados - rechazados
        tasa_aprobacion = round((aprobados / validados) * 100, 1)

    # Gráficos filtrados
    estados_data = reclamos.values('estado').annotate(total=Count('id'))
    estados_labels = [item['estado'].replace('_', ' ').title() for item in estados_data]
    estados_counts = [item['total'] for item in estados_data]

    areas_data = AsignacionArea.objects.filter(
        reclamo__in=reclamos
    ).values('departamento').annotate(total=Count('reclamo', distinct=True))
    areas_labels = [item['departamento'] for item in areas_data]
    areas_counts = [item['total'] for item in areas_data]

    # Ranking de clientes filtrado
    ranking_clientes = reclamos.values(
        'cliente__user__first_name',
        'cliente__user__last_name',
        'cliente__company_name'
    ).annotate(
        total=Count('id'),
        resueltos=Count('id', filter=Q(estado__in=['resuelto', 'cerrado'])),
        pendientes=Count('id', filter=Q(estado__in=['pendiente_validacion', 'en_proceso', 'resolucion_propuesta'])),
    ).order_by('-total')[:10]

    aprobados_count = reclamos.filter(estado__in=['en_proceso', 'resolucion_propuesta', 'resuelto', 'cerrado']).count()
    rechazados_count = rechazados

    return render(request, 'reclamos/panel_gerencia.html', {
        'reclamos': reclamos,
        'total_reclamos': total_reclamos,
        'pendientes_calidad': pendientes_calidad,
        'aprobados_count': aprobados_count,
        'resueltos': resueltos,
        'rechazados': rechazados,
        'rechazados_count': rechazados_count,
        'tiempo_promedio_dias': tiempo_promedio_dias,
        'tasa_aprobacion': tasa_aprobacion,
        'estados_labels': json.dumps(estados_labels),
        'estados_counts': json.dumps(estados_counts),
        'areas_labels': json.dumps(areas_labels),
        'areas_counts': json.dumps(areas_counts),
        'ranking_clientes': ranking_clientes,
        'periodo': periodo,
        'fecha_inicio_str': fecha_inicio_str,
        'fecha_fin_str': fecha_fin_str,
    })

# ==========================================
# EXPORTAR EXCEL
# ==========================================
@login_required
def exportar_excel(request):
    if not request.user.is_superuser:
        return redirect('login')

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Reclamos Alexandra Farms'

    # Estilo encabezado
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1a3c34', end_color='1a3c34', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')

    # Encabezados
    headers = [
        'Número', 'Cliente', 'Empresa', 'País',
        'Tipo de Problema', 'Variedad', 'Estado',
        'Inspector Calidad', 'Fecha Creación', 'Última Actualización'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Ancho de columnas
    anchos = [20, 25, 25, 15, 30, 15, 25, 25, 20, 20]
    for col, ancho in enumerate(anchos, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho

    # Datos
    reclamos = Reclamo.objects.select_related(
        'cliente__user', 'inspector_calidad__user', 'tipo_problema'
    ).order_by('-created_at')

    for row, r in enumerate(reclamos, 2):
        ws.cell(row=row, column=1, value=r.numero)
        ws.cell(row=row, column=2, value=r.cliente.user.get_full_name())
        ws.cell(row=row, column=3, value=r.cliente.company_name)
        ws.cell(row=row, column=4, value=r.cliente.country)
        ws.cell(row=row, column=5, value=str(r.tipo_problema))
        ws.cell(row=row, column=6, value=r.variedad_rosa)
        ws.cell(row=row, column=7, value=r.get_estado_display())
        ws.cell(row=row, column=8, value=r.inspector_calidad.user.get_full_name() if r.inspector_calidad else '--')
        ws.cell(row=row, column=9, value=r.created_at.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row, column=10, value=r.updated_at.strftime('%d/%m/%Y %H:%M') if r.updated_at else '--')

        # Filas alternadas
        if row % 2 == 0:
            fill = PatternFill(start_color='e8f5f0', end_color='e8f5f0', fill_type='solid')
            for col in range(1, 11):
                ws.cell(row=row, column=col).fill = fill

    # Respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="reclamos_alexandra_farms.xlsx"'
    wb.save(response)
    return response


# ==========================================
# EXPORTAR PDF
# ==========================================
@login_required
def exportar_pdf(request):
    if not request.user.is_superuser:
        return redirect('login')

    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from django.http import HttpResponse
    from django.utils import timezone

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reclamos_alexandra_farms.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        rightMargin=1*cm,
        leftMargin=1*cm,
        topMargin=1.5*cm,
        bottomMargin=1*cm
    )

    elementos = []
    styles = getSampleStyleSheet()

    # Título
    titulo_style = ParagraphStyle(
        'titulo',
        parent=styles['Title'],
        fontSize=16,
        textColor=colors.HexColor('#1a3c34'),
        spaceAfter=6,
    )
    subtitulo_style = ParagraphStyle(
        'subtitulo',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#757575'),
        spaceAfter=20,
    )

    elementos.append(Paragraph('Alexandra Farms — Reporte de Reclamos', titulo_style))
    elementos.append(Paragraph(f'Generado el {timezone.now().strftime("%d/%m/%Y %H:%M")}', subtitulo_style))

    # Tabla
    reclamos = Reclamo.objects.select_related(
        'cliente__user', 'inspector_calidad__user', 'tipo_problema'
    ).order_by('-created_at')

    data = [['Número', 'Cliente', 'Empresa', 'Problema', 'Estado', 'Inspector', 'Fecha']]

    for r in reclamos:
        data.append([
            r.numero,
            r.cliente.user.get_full_name(),
            r.cliente.company_name,
            str(r.tipo_problema)[:30],
            r.get_estado_display(),
            r.inspector_calidad.user.get_full_name() if r.inspector_calidad else '--',
            r.created_at.strftime('%d/%m/%Y'),
        ])

    tabla = Table(data, colWidths=[3.5*cm, 4*cm, 4*cm, 6*cm, 4*cm, 4*cm, 3*cm])
    tabla.setStyle(TableStyle([
        # Encabezado
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a3c34')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        ('ROWBACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a3c34')),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        # Filas
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        # Filas alternadas
        ('ROWBACKGROUND', (0, 1), (-1, -1), colors.white),
        ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#e8f5f0')),
        ('BACKGROUND', (0, 4), (-1, 4), colors.HexColor('#e8f5f0')),
        ('BACKGROUND', (0, 6), (-1, 6), colors.HexColor('#e8f5f0')),
        ('BACKGROUND', (0, 8), (-1, 8), colors.HexColor('#e8f5f0')),
        ('BACKGROUND', (0, 10), (-1, 10), colors.HexColor('#e8f5f0')),
        # Bordes
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e0e0e0')),
        ('LINEBELOW', (0, 0), (-1, 0), 1, colors.HexColor('#1a3c34')),
    ]))

    elementos.append(tabla)
    doc.build(elementos)
    return response